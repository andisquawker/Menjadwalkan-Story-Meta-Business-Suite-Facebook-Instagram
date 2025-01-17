from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from datetime import timedelta
from datetime import datetime
import pyautogui
import time
import random
import pyperclip
from openpyxl import load_workbook

# Mengatur waktu tunggu acak antara 1 hingga 5 detik
random_wait = random.uniform(1, 5)

# Menentukan jumlah hari maksimal yang diizinkan untuk penjadwalan
max_days_allowed = 28

# Mengatur opsi untuk Chrome WebDriver
opsi = webdriver.ChromeOptions()
opsi.add_experimental_option("detach", True)
opsi.add_argument("--disable-notifications")
opsi.add_argument("--mute-audio")
opsi.add_argument("--disable-blink-features=AutomationControlled")
opsi.add_experimental_option("excludeSwitches", ["enable-automation"])
opsi.add_argument("--user-data-dir=Ganti dengan direktori profil Chrome Anda")

# ID Fanspage Hiburan Duniawi
Nama_Fanspage = "Id Fanspage"

# Nama file Excel yang digunakan
file_excel = "File_Excel.xlsx"
id_fanspage = Nama_Fanspage

# Memuat workbook dan worksheet dari file Excel
wb = load_workbook(file_excel)
ws = wb.active

# Memulai WebDriver Chrome
driver = webdriver.Chrome(options=opsi)
driver.window_handles
driver.get("https://www.facebook.com/")

# Menunggu hingga halaman Facebook terbuka
WebDriverWait(driver, 30).until(EC.url_contains("facebook.com"))

# Iterasi melalui baris-baris di worksheet Excel
for idx, baris in enumerate(
    ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2
):
    laporan = baris[0]
    Source_Media = baris[1]
    caption = baris[2]
    jadwaltayang = baris[3]
    jam = baris[4]
    menit = baris[5]

    # Jika laporan sudah di-upload, lewati baris ini
    if laporan == "Upload":
        print(f"Skip Video {caption} Karena Laporan Sudah Terupload")
        continue

    # Jika data tidak lengkap, tutup driver dan hentikan program
    if not all([Source_Media, caption, jadwaltayang, jam, menit]):
        print(f"Data pada baris {idx} tidak lengkap. Menutup driver.")
        driver.quit()
        break

    # Memeriksa format tanggal
    try:
        jadwal_tayang_date = datetime.strptime(jadwaltayang, "%d/%m/%Y")
    except ValueError:
        print(f"Format tanggal tidak valid di baris {idx}. Menutup driver.")
        driver.quit()
        break

    # Memeriksa apakah tanggal tayang lebih dari 28 hari ke depan
    if jadwal_tayang_date > datetime.today() + timedelta(days=max_days_allowed):
        print(
            f"Tanggal Tayang Di Baris {idx} Lebih Dari 28 Hari Ke Depan. Menutup Driver."
        )
        driver.quit()
        break

    # Membuka halaman Facebook Business untuk fanspage
    driver.get(f"https://business.facebook.com/latest/home?asset_id={id_fanspage}")
    WebDriverWait(driver, 30).until(EC.url_contains("business.facebook.com"))

    # Membuka halaman Story Composer
    driver.get(
        "https://business.facebook.com/latest/story_composer/?ref=biz_web_home_stories&context_ref=HOME"
    )
    WebDriverWait(driver, 30).until(EC.url_contains("story_composer"))

    # Menunggu hingga elemen untuk mengunggah media dapat diklik
    WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable(
            (
                By.XPATH,
                "//div[@class='x1i10hfl xjqpnuy xa49m3k xqeqjp1 x2hbi6w x972fbf xcfux6l x1qhh985 xm0m39n x9f619 x1ypdohk xe8uvvx xdj266r x11i5rnm xat24cr x1mh8g0r x2lwn1j xeuugli x16tdsg8 xggy1nq x1ja2u2z x6s0dn4 x1ejq31n xd10rxx x1sy0etr x17r0tee x3nfvp2 xdl72j9 x1q0g3np x2lah0s x193iq5w x1n2onr6 x1hl2dhg x87ps6o xxymvpz xlh3980 xvmahel x1lku1pv x1g40iwv x1g2r6go x16e9yqp x12w9bfk x15406qy x1lcm9me x1yr5g0i xrt01vj x10y3i5r x1ob88yx xaatb59 x1qgsegg xo1l8bm xbsr9hj x1v911su x1y1aw1k xwib8y2 x1pi30zi x1ye3gou']",
            )
        )
    ).click()

    time.sleep(3)

    # Menyalin dan menempelkan sumber media menggunakan pyautogui dan pyperclip
    pyperclip.copy(str(Source_Media))
    pyautogui.hotkey("ctrl", "v")
    pyautogui.press("enter")

    driver.implicitly_wait(60)
    Indikator = driver.find_element(
        By.XPATH, "//div[@class='_67wx']/img[1] | //video[@class='_67wy']"
    )

    driver.implicitly_wait(30)
    Opsi_Penjadwalan = driver.find_element(
        By.XPATH,
        "//div[@class='x6s0dn4 x78zum5 x1q0g3np x1qughib x2lwn1j xeuugli']//div[@class='x1i10hfl xjqpnuy xa49m3k xqeqjp1 x2hbi6w x972fbf xcfux6l x1qhh985 xm0m39n x9f619 x1ypdohk xe8uvvx xdj266r x11i5rnm xat24cr x1mh8g0r x2lwn1j xeuugli x16tdsg8 xggy1nq x1ja2u2z x6s0dn4 x1ejq31n xd10rxx x1sy0etr x17r0tee x3nfvp2 xdl72j9 x1q0g3np x2lah0s x193iq5w x1n2onr6 x1hl2dhg x87ps6o xxymvpz xlh3980 xvmahel x1lku1pv x1g40iwv x1g2r6go x16e9yqp x12w9bfk x15406qy x1lcm9me x1yr5g0i xrt01vj x10y3i5r x1ob88yx xaatb59 x1qgsegg xo1l8bm xbsr9hj x1v911su x1y1aw1k xwib8y2 x1swvt13 x1pi30zi']",
    )
    Opsi_Penjadwalan.click()

    # Mengisi tanggal tayang
    tanggaltayang_elements = driver.find_elements(
        By.CSS_SELECTOR, "[placeholder='dd/mm/yyyy']"
    )

    if len(tanggaltayang_elements) >= 1:

        tanggaltayang_elements[0].send_keys(Keys.CONTROL, "a")
        tanggaltayang_elements[0].send_keys(jadwaltayang)
        time.sleep(3)

        if len(tanggaltayang_elements) >= 2:

            tanggaltayang_elements[1].send_keys(Keys.CONTROL, "a")
            tanggaltayang_elements[1].send_keys(jadwaltayang)
            time.sleep(3)
    else:
        print("Tidak ada cukup elemen tanggal tayang ditemukan")

    driver.implicitly_wait(60)

    # Mengisi jam tayang
    jamtayang_elements = driver.find_elements(By.CSS_SELECTOR, "[aria-label='jam']")
    menittayang_elements = driver.find_elements(By.CSS_SELECTOR, "[aria-label='menit']")

    if len(jamtayang_elements) >= 1:
        jamtayang_elements[0].send_keys(jam)

        if len(jamtayang_elements) >= 2:
            jamtayang_elements[1].send_keys(jam)
    else:
        print("Tidak ada cukup elemen jam tayang ditemukan")

    # Mengisi menit tayang
    if len(menittayang_elements) >= 1:
        menittayang_elements[0].send_keys(menit)

        if len(menittayang_elements) >= 2:
            menittayang_elements[1].send_keys(menit)
    else:
        print("Tidak ada cukup elemen menit tayang ditemukan")

    time.sleep(5)

    # Menjadwalkan story
    jadwalkan = driver.find_element(
        By.XPATH,
        "//div[@class='x1i10hfl xjqpnuy xa49m3k xqeqjp1 x2hbi6w x972fbf xcfux6l x1qhh985 xm0m39n x9f619 x1ypdohk xe8uvvx xdj266r x11i5rnm xat24cr x1mh8g0r x2lwn1j xeuugli x16tdsg8 xggy1nq x1ja2u2z x1t137rt x6s0dn4 x1ejq31n xd10rxx x1sy0etr x17r0tee x3nfvp2 xdl72j9 x1q0g3np x2lah0s x193iq5w x1n2onr6 x1hl2dhg x87ps6o xxymvpz xlh3980 xvmahel x1lku1pv x1g40iwv x1g2r6go x16e9yqp x12w9bfk x15406qy x1lcm9me x1yr5g0i xrt01vj x10y3i5r xo1l8bm x140t73q x19bke7z x1y1aw1k xwib8y2 x1swvt13 x1pi30zi']",
    )
    jadwalkan.click()
    time.sleep(10)

    # Mengupdate status laporan di Excel
    laporan = "Upload".format(idx)
    ws.cell(row=idx, column=1, value=laporan)
    wb.save(file_excel)

    print(f"Selesai Menjadwalkan Story {caption}")

    time.sleep(10)

print("Selesai Penjadwalan Reel Facebook")

# Menutup driver
driver.quit()
